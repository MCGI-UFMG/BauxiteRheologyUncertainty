# -*- coding: utf-8 -*-
"""
Created on Tue Feb 20 14:26:48 2024

@author: rodri
"""

import h5py
import pandas as pd
import numpy as np
import openpyxl
import glob, os

#ATENÇÃO: ATUALIZAR PRÓXIMA LINHA COM CAMINHO DA PASTA DE RESULTADOS HDF
folder="D:/RAS_results/RAS3/"
os.chdir(folder)

#ATENÇÃO: ATUALIZAR PRÓXIMA LINHA COM CAMINHO DO ARQUIVO id_cells.txt
cells_ID=pd.read_csv(r'D:/RAS_results/id_cells.txt',sep="\t")

#ATENÇÃO: ATUALIZAR PRÓXIMA LINHA COM CAMINHO DO ARQUIVO id_face.txt
face_ID=pd.read_csv(r'D:/RAS_results/id_face.txt',sep="\t")

arquivos = glob.glob("*.hdf")


#Extraindo resultados de Velocidade
main_vel_list = [pd.DataFrame() for _ in range(5)]

j=1
for file in arquivos:
    with h5py.File(folder+file,'r') as hdf:
        data=np.array(hdf.get('Results/Unsteady/Output/Output Blocks/Base Output/Unsteady Time Series/2D Flow Areas/Perimeter 1/Face Velocity'))
        vel_max=pd.DataFrame(np.max(np.array(data),axis=0))
        vel_max.index=vel_max.index+1
        k=0
        for i in face_ID.columns:
            index=list(face_ID[i])
            index = [x - 1 for x in index]
            main_vel_list[k][j]=vel_max.iloc[index]
            k=k+1
            print(k)
    j=j+1

#Extraindo resultados de WSE
main_wse_list = [pd.DataFrame() for _ in range(5)]

j=1
for file in arquivos:
    with h5py.File(folder+file,'r') as hdf:
        data=np.array(hdf.get('Results/Unsteady/Output/Output Blocks/Base Output/Unsteady Time Series/2D Flow Areas/Perimeter 1/Water Surface'))
        wse_max=pd.DataFrame(np.max(np.array(data),axis=0))
        wse_max.index=wse_max.index+1
        k=0
        for i in cells_ID.columns:
            index=list(cells_ID[i])
            index = [x - 1 for x in index]
            main_wse_list[k][j]=wse_max.iloc[index]
            k=k+1
            print(k)
    j=j+1

#Extraindo resultados de Minimum Elevation
main_minEl_list = [pd.DataFrame() for _ in range(5)]

file = arquivos[0]

j=1
for file in arquivos:
    with h5py.File(folder+file,'r') as hdf:
        data = np.array(hdf.get('Geometry/2D Flow Areas/Perimeter 1/Cells Minimum Elevation'))
        k=0
        for i in cells_ID.columns:
            index = list(cells_ID[i])
            index = [x for x in index]
            main_minEl_list[k][j] = pd.DataFrame(data[index], index=index)
            k=k+1
    j=j+1

#Salvando resultados de Velocidade

#ATENÇÃO: ATUALIZAR PRÓXIMA LINHA COM CAMINHO QUE DESEJA SALVAR OS RESULTADOS DE VELOCIDADE
with pd.ExcelWriter(r'D:/RAS_results/RAS3/vel_output.xlsx') as writer: 
    i=1
    for x in main_vel_list: 
        x.to_excel(writer, sheet_name="XS"+str(i), index=False) 
        i=i+1

#Salvando resultados de WSE

#ATENÇÃO: ATUALIZAR PRÓXIMA LINHA COM CAMINHO QUE DESEJA SALVAR OS RESULTADOS DE WSE
with pd.ExcelWriter(r'D:/RAS_results/RAS3/wse_output.xlsx') as writer: 
    i=1
    for x in main_wse_list: 
        x.to_excel(writer, sheet_name="XS"+str(i), index=False) 
        i=i+1

#Salvando resultados de WSE

#ATENÇÃO: ATUALIZAR PRÓXIMA LINHA COM CAMINHO QUE DESEJA SALVAR OS RESULTADOS DE MIN ELEVATION
with pd.ExcelWriter(r'D:/RAS_results/RAS3/min-El_output.xlsx') as writer: 
    i=1
    for x in main_minEl_list: 
        x.to_excel(writer, sheet_name="XS"+str(i), index=False) 
        i=i+1


# Inserir linhas vazias
excel_files = ['vel_output.xlsx', 'wse_output.xlsx', 'min-El_output.xlsx']

for file in excel_files:
    caminho_completo = os.path.join(folder, file)
    workbook = openpyxl.load_workbook(caminho_completo)
    
    # Iterar sobre todas as planilhas do arquivo
    for sheet in workbook.sheetnames:
        workbook[sheet].insert_rows(2)
    
    workbook.save(caminho_completo)



# Função para extrair a string gXXuYY da primeira linha de cada arquivo
def extrair_string_primeira_linha(arquivo):
    with open(arquivo, 'r') as file:
        primeira_linha = file.readline().strip()
        _, string_gXXuYY = primeira_linha.split('=')
    return string_gXXuYY

# Lendo os arquivos .pZZ ordenados
arquivos_pZZ = sorted(glob.glob(os.path.join(folder, '*.p??')))

# Extrair a string gXXuYY de cada arquivo
strings_gXXuYY = [extrair_string_primeira_linha(arquivo) for arquivo in arquivos_pZZ]

# Função para atualizar os arquivos Excel
def atualizar_excel(arquivo_excel, strings_gXXuYY):
    workbook = openpyxl.load_workbook(arquivo_excel)
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for col, valor in enumerate(strings_gXXuYY, start=1):
            ws.cell(row=2, column=col).value = valor
    workbook.save(arquivo_excel)

# Atualizar os arquivos Excel
atualizar_excel('vel_output.xlsx', strings_gXXuYY)
atualizar_excel('wse_output.xlsx', strings_gXXuYY)
atualizar_excel('min-El_output.xlsx', strings_gXXuYY)
